#!/usr/bin/python3
# -*- coding: utf-8 -*-
# Author: Valdemir Bezerra

import openpyxl
import os
from datetime import date
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook('salvatroughput.xlsx')

roteadores = ['JPADTCRTD05.sh', 'JPADTCRTD01.sh', 'JPAMIRRTD01.sh', 'CGEDTCRTD01.sh']

for roteador in roteadores:

    sheet = wb.get_sheet_by_name(roteador)

    hoje = date.today()
#for row in sheet.iter_rows(min_row=1, max_col=2, max_row=4):
#   for cell in row:
#        print(cell) #cell.value

#print("Por colunas")

#for col in sheet.iter_cols(min_row=1, max_col=2, max_row=3):
#    for cell in col:
#        print(cell)
    ultimaColunaUsada = sheet.max_column
    novaColuna = ultimaColunaUsada + 1
    ultimaLinhaUsada = sheet.max_row


    sheet.cell(row=1, column=novaColuna).fill = PatternFill(fill_type='solid', start_color='EE1111',end_color='EE1111')

    sheet.cell(row=1, column=novaColuna + 1).fill = PatternFill(fill_type='solid', start_color='EE1111',end_color='EE1111')



    sheet.cell(row=1, column=novaColuna).value = "MEDIA-in-" + str(hoje.day) + "/" + str(hoje.month) + "/" + str(
            hoje.year)[2:4]
    sheet.cell(row=1, column=novaColuna + 1).value = "MEDIA-out-" + str(hoje.day) + "/" + str(
            hoje.month) + "/" + str(hoje.year)[2:4]

    score = 2

    scoreLinhas = score * 2



    for linhaSelecionada in range(2, ultimaLinhaUsada + 1):

        coordenadasLinhasIn = " "
        coordenadasLinhasOut = " "
        dividendo = 0

        for linhaMedia in range(scoreLinhas, 0, -1):
        #print(linhaMedia) # funcao decrescente pra pegar as celular voltando e juntar numa string so
            if dividendo % 2 == 0:
                coordenadasLinhasIn += sheet.cell(row=linhaSelecionada, column=novaColuna - linhaMedia).coordinate + "+"

            else:
                coordenadasLinhasOut += sheet.cell(row=linhaSelecionada, column=novaColuna - linhaMedia).coordinate + "+"


            dividendo +=1

    #print(coordenadasLinhasIn[: -1 ])
    #print(coordenadasLinhasOut[: -1 ])

        sheet.cell(row=linhaSelecionada, column=novaColuna).value = '=(('+ coordenadasLinhasIn[: -1 ] + ')/'+ str(score) + ')/1000000000'

        sheet.cell(row=linhaSelecionada, column=novaColuna + 1).value = '=(('+ coordenadasLinhasOut[: -1 ] + ')/'+ str(score) + ')/1000000000'


    wb.save('salvatroughput.xlsx')