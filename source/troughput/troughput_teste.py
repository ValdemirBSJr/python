#!/usr/bin/python3
# -*- coding: utf-8 -*-
# Author: Valdemir Bezerra

import os, sys
import subprocess
from subprocess import Popen, PIPE
import openpyxl
import re
from datetime import date


# Se nao for o root, gera excecao
#if not os.geteuid()==0:
#    mensagemErro = "\nErro:\n Voce deve estar logado como root no prompt de comando para usar essa aplicaçao.\nUsuario ativo: " + os.environ['USER']
#    sys.exit(mensagemErro)

print("Acesso root concedido! Dando inicio a coleta de troughput...")


#p = subprocess.Popen(["echo", "hello world"], stdout=subprocess.PIPE)

#print(p.communicate())

#p1 = subprocess.Popen(['ping', '-c 2', '8.8.8.8'], stdout=subprocess.PIPE)


# Run the command
#output = p1.communicate()[0]

#print(output)

#print(os.system('echo $HOME'))

#processo = subprocess.run('./JPADTCRTD01.sh', shell=True)
#processo = subprocess.run('./JPADTCRTD05.sh', shell=True)
#processo = subprocess.run('./JPAMIRRTD01.sh', shell=True)
#processo = subprocess.run('./teste_TCL.sh')

#Abaixo huawei

retornoHW = ''''<JPADTCRTD01>display interface giga 6/0/0 | in Last 300
    Last 300 seconds input rate: 214725824 bits/sec, 80648 packets/sec
    Last 300 seconds output rate: 1659786184 bits/sec, 165284 packets/sec'''


#retorno2 = retornoHW.split(' ')
#print("Input: ", retorno2[16])
#print("Output: ", retorno2[28])

retornoCisco = """RP/0/RSP0/CPU0:JPADTCRTD05#show interface Te0/3/0/6 | in 30 second
Sun Nov 12 01:27:30.231 GMT
  30 second input rate 2719776000 bits/sec, 298350 packets/sec
  30 second output rate 831415000 bits/sec, 170399 packets/sec
"""

pegaregex = re.compile(r'(\d+)\sbits/sec') #Pega exatamente os bites seguidos de "bits/sec"
#print(pegaregex.findall(retornoHW))
#print(pegaregex.findall(retornoCisco))

conteudoArquivo = open(os.path.join('./tmp/', 'JPADTCRTD05')) #junta as pastas por ordem para fazer o caminho

stringArquivo = conteudoArquivo.read() #Aqui eu leio todo o arquivo como se fosse uma string

listaTroughput= pegaregex.findall(stringArquivo)

print(listaTroughput)

wb = openpyxl.load_workbook('salvatroughput.xlsx')
#print(type(wb))
sheet = wb.get_sheet_by_name('JPADTCRTD05')
#sheet['A1'] = retorno2[16]
#sheet['B1'] = retorno3[16]

ultimaColunaUsada = sheet.max_column
ultimaLinhaUsada = sheet.max_row + 1

#get_highest estao depreciados agora se usa max_column e max_row

#print(str(ultimaColunaUsada))
#print(str(ultimaLinhaUsada))

#sheet.columns[ultimaColunaUsada]

############## DATAS ##########################

#vou pegar o dia pra por no titulo da coluna, se for sexta ele acrescenta mais uma e faz o calculo

hoje = date.today()
#print(hoje.weekday())
# 4 e sexta

############### FIM DATAS ########################
novaColuna = ultimaColunaUsada + 1

sheet.cell(row=1, column=novaColuna).value = "data-in-" + str(hoje.day) + "/" + str(hoje.month) + "/" + str(hoje.year)[2:4]
sheet.cell(row=1, column=novaColuna + 1).value = "data-out-" + str(hoje.day) + "/" + str(hoje.month) + "/" + str(hoje.year)[2:4]

linhaInicialIn = 2
linhaInicialOut = 2
dividendo = 0
#print(len(listaTroughput))

for i in listaTroughput:
#Nesse caso voce deve usar o proprio i ao inves de tentar referenciar ele em listaTroughput[i]
# o i e o proprio elemento e nao indica iteracao

    #indice = listaTroughput.index(i)
    #print(i)


    if dividendo % 2 == 0:
        sheet.cell(row=linhaInicialIn, column=novaColuna).value = i
        linhaInicialIn += 1
    else:
        sheet.cell(row=linhaInicialOut, column=novaColuna + 1).value = i
        linhaInicialOut += 1

    dividendo += 1

wb.save('salvatroughput.xlsx')