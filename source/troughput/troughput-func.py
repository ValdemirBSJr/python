#!/usr/bin/python3
# -*- coding: utf-8 -*-
# Author: Valdemir Bezerra

import os, sys
import subprocess
import openpyxl
from openpyxl.styles import PatternFill
import re
from datetime import date
import shelve


print('''

  _______   _____     ____    _    _    _____   _    _   _____    _    _   _______ 
 |__   __| |  __ \   / __ \  | |  | |  / ____| | |  | | |  __ \  | |  | | |__   __|
    | |    | |__) | | |  | | | |  | | | |  __  | |__| | | |__) | | |  | |    | |   
    | |    |  _  /  | |  | | | |  | | | | |_ | |  __  | |  ___/  | |  | |    | |   
    | |    | | \ \  | |__| | | |__| | | |__| | | |  | | | |      | |__| |    | |   
    |_|    |_|  \_\  \____/   \____/   \_____| |_|  |_| |_|       \____/     |_|   
                                                                                   
                                                                                   
                        Powered on PYTHON!!!
                        Developer: N5669203

''')



#if not os.geteuid()==0:
#    mensagemErro = "\nErro:\n Voce deve estar logado como root no prompt de comando para usar essa aplicaçao.\nUsuario ativo: " + os.environ['USER']
#    sys.exit(mensagemErro)

print("Acesso root concedido! Dando inicio a coleta de troughput...")

def salvaExcel(**kwargs):

    hoje = date.today()

    wb = openpyxl.load_workbook('./equissel/salvatroughput.xlsx', data_only=True)
    sheet = wb.get_sheet_by_name(kwargs['roteador'])

    ultimaColunaUsada = sheet.max_column
    ultimaLinhaUsada = sheet.max_row + 1

    novaColuna = ultimaColunaUsada + 1

    sheet.cell(row=1, column=novaColuna).value = "data-in-" + str(hoje.day) + "/" + str(hoje.month) + "/" + str(
            hoje.year)[2:4]
    sheet.cell(row=1, column=novaColuna + 1).value = "data-out-" + str(hoje.day) + "/" + str(
            hoje.month) + "/" + str(hoje.year)[2:4]

    linhaInicialIn = 2
    linhaInicialOut = 2
    dividendo = 0

    scoreLinhas = kwargs['escore'] * 2

    for i in kwargs['listaT']:
        if dividendo % 2 == 0:
            sheet.cell(row=linhaInicialIn, column=novaColuna).value = i
            linhaInicialIn += 1
        else:
            sheet.cell(row=linhaInicialOut, column=novaColuna + 1).value = i
            linhaInicialOut += 1

        dividendo += 1

#########################################################################################################

    if kwargs['escore'] > 0:

        print("Hoje e sexta-feira, dia de maldade e de calcular a media semanal do troughput.")
        print("Calculando...")

        ultimaColunaUsada = sheet.max_column
        novaColuna = ultimaColunaUsada + 1
        ultimaLinhaUsada = sheet.max_row

        sheet.cell(row=1, column=novaColuna).fill = PatternFill(fill_type='solid', start_color='EE1111',end_color='EE1111')

        sheet.cell(row=1, column=novaColuna + 1).fill = PatternFill(fill_type='solid', start_color='EE1111',end_color='EE1111')



        sheet.cell(row=1, column=novaColuna).value = "MEDIA-in-" + str(hoje.day) + "/" + str(hoje.month) + "/" + str(
            hoje.year)[2:4]
        sheet.cell(row=1, column=novaColuna + 1).value = "MEDIA-out-" + str(hoje.day) + "/" + str(
            hoje.month) + "/" + str(hoje.year)[2:4]



        for linhaSelecionada in range(2, ultimaLinhaUsada + 1):

            coordenadasLinhasIn = " "
            coordenadasLinhasOut = " "
            dividendoLinha = 0

            for linhaMedia in range(scoreLinhas, 0, -1):
                # print(linhaMedia) # funcao decrescente pra pegar as celular voltando e juntar numa string so
                if dividendo % 2 == 0:
                    coordenadasLinhasIn += sheet.cell(row=linhaSelecionada, column=novaColuna - linhaMedia).coordinate + "+"

                else:
                    coordenadasLinhasOut += sheet.cell(row=linhaSelecionada, column=novaColuna - linhaMedia).coordinate + "+"

                dividendoLinha += 1



            sheet.cell(row=linhaSelecionada, column=novaColuna).value = '=((' + coordenadasLinhasIn[: -1] + ')/' + str(kwargs['escore']) + ')/1000000000'

            sheet.cell(row=linhaSelecionada, column=novaColuna + 1).value = '=((' + coordenadasLinhasOut[: -1] + ')/' + str(kwargs['escore']) + ')/1000000000'




#############################################################################################################
    wb.save('./equissel/salvatroughput.xlsx')
    print('Arquivo excel atualizado com sucesso!')


#####################################################

def calculaMediaExcel(**kwargs):

    hoje = date.today()

    wb = openpyxl.load_workbook('./equissel/salvatroughput.xlsx', data_only=True)
    sheet = wb.get_sheet_by_name(kwargs['roteador'])

    ultimaColunaUsada = sheet.max_column
    novaColuna = ultimaColunaUsada + 1
    ultimaLinhaUsada = sheet.max_row

    sheet.cell(row=1, column=novaColuna).fill = PatternFill(fill_type='solid', start_color='EE1111', end_color='EE1111')

    sheet.cell(row=1, column=novaColuna + 1).fill = PatternFill(fill_type='solid', start_color='EE1111',
                                                                end_color='EE1111')

    sheet.cell(row=1, column=novaColuna).value = "MEDIA-in-" + str(hoje.day) + "/" + str(hoje.month) + "/" + str(
        hoje.year)[2:4]
    sheet.cell(row=1, column=novaColuna + 1).value = "MEDIA-out-" + str(hoje.day) + "/" + str(
        hoje.month) + "/" + str(hoje.year)[2:4]

    scoreLinhas = kwargs['escore'] * 2

    for linhaSelecionada in range(2, ultimaLinhaUsada + 1):

        coordenadasLinhasIn = " "
        coordenadasLinhasOut = " "
        dividendoLinha = 0

        for linhaMedia in range(scoreLinhas, 0, -1):
        # print(linhaMedia) # funcao decrescente pra pegar as celular voltando e juntar numa string so
            if dividendoLinha % 2 == 0:
                coordenadasLinhasIn += sheet.cell(row=linhaSelecionada, column=novaColuna - linhaMedia).coordinate + "+"

            else:
                coordenadasLinhasOut += sheet.cell(row=linhaSelecionada, column=novaColuna - linhaMedia).coordinate + "+"

            dividendoLinha += 1

        sheet.cell(row=linhaSelecionada, column=novaColuna).value = '=((' + coordenadasLinhasIn[: -1] + ')/' + str(
        kwargs['escore']) + ')/1000000000'

        sheet.cell(row=linhaSelecionada, column=novaColuna + 1).value = '=((' + coordenadasLinhasOut[: -1] + ')/' + str(
        kwargs['escore']) + ')/1000000000'

    wb.save('./equissel/salvatroughput.xlsx')


#####################################################

def salvaEscore(somaEscore):

    if os.path.exists(os.path.join('./scr/', 'score')) == True:

        if somaEscore == 0:

            arquivoSRC = shelve.open(os.path.join('./scr/', 'score'))

            valor = arquivoSRC['scr']

            soma = int(valor) + 1

            arquivoSRC['scr'] = soma

            arquivoSRC.close()

        elif somaEscore == 1:
            arquivoSRC = shelve.open(os.path.join('./teste_shelve/', 'score'))
            scr = 0
            arquivoSRC['scr'] = scr
            arquivoSRC.close()

    else:

        arquivoSRC = shelve.open(os.path.join('./teste_shelve/', 'score'))
        scr = 0
        arquivoSRC['scr'] = scr
        arquivoSRC.close()



def exec_rotina(*args):
    hoje = date.today()

    if hoje.weekday() == 4:
        salvaEscore(0)

        arquivoSRC = shelve.open(os.path.join('./scr/', 'score'))
        valor = arquivoSRC['scr']
        arquivoSRC.close()

        for roteador in args:
            processo = subprocess.run(os.path.join('./rt/', roteador), shell=True)

            pegaregex = re.compile(r'(\d+)\sbits/sec')

            arquivoSalvo = roteador
            retornoRoteador = arquivoSalvo.split(".")

            conteudoArquivo = open(os.path.join('./tmp/', retornoRoteador[0]))
            stringArquivo = conteudoArquivo.read()

            listaTroughput = pegaregex.findall(stringArquivo)

            salvaExcel(roteador=arquivoSalvo, escore=valor, listaT=listaTroughput)
##########################################################################
            calculaMediaExcel(roteador=arquivoSalvo, escore=valor)
###########################################################################
        salvaEscore(1)

    else:

        salvaEscore(0)

        for roteador in args:
            processo = subprocess.run(os.path.join('./rt/', roteador), shell=True)

            pegaregex = re.compile(r'(\d+)\sbits/sec')

            arquivoSalvo = roteador
            retornoRoteador = arquivoSalvo.split(".")

            conteudoArquivo = open(os.path.join('./tmp/', retornoRoteador[0]))
            stringArquivo = conteudoArquivo.read()

            listaTroughput = pegaregex.findall(stringArquivo)

            salvaExcel(roteador=arquivoSalvo, escore=0, listaT=listaTroughput)


print('Listando roteadores para consulta...')

roteadores = os.listdir(os.path.join('./rt/'))

exec_rotina(*roteadores)

for i in roteadores:

    arquivoSalvo = i
    retornoRoteador = arquivoSalvo.split(".")

    os.remove(os.path.join('./tmp/', retornoRoteador[0]))


print("")
print("Coleta de troughput realizada com sucesso!")

