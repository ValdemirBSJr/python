#!/home/user/.virtualenvs/k36/bin/python
# -*- coding: utf-8 -*-
# Author: Valdemir Bezerra


# Imports

import requests
import bs4
from requests.auth import HTTPBasicAuth
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import os
import subprocess
from datetime import date
import re
import sys
import click



hoje = date.today()
diaTratativa = date.fromordinal(hoje.toordinal() - 1)
mesTratativa = str(diaTratativa.month)
anoTratativa = str(diaTratativa.year)


def verificaCaminho(caminhoArquivo):
    if os.path.isfile(caminhoArquivo) is True:
        return 'valido'
    else:
        return 'invalido'


def lerLista(**kwargs):
    lidoMemoria = open(kwargs['lista'])
    listaMontada = lidoMemoria.readlines()

    # abaixo temos uma forma de retirar todos os : dos MACs
    listaMontada = ''.join(listaMontada).replace(':', '').split()
    print(' ')
    print('=' * 70)
    print('Pronto! Agora vamos consultar os contratos, isto pode demorar vários minutos...')

    return listaMontada


def fazConsulta(**kwargs):
    caminho = '/home/user/Documentos/SCRIPTS/consultaPEV/saida/consultados.xlsx'


    if verificaCaminho(caminho) == 'valido':

        nomeSheet = str(diaTratativa.day) + "." + mesTratativa + "." + anoTratativa
        wb = load_workbook(filename=caminho)
        wsdia = wb.create_sheet(nomeSheet)
        wb[nomeSheet]

        wsdia['A1'] = 'CONTRATO'
        wsdia['B1'] = 'CM-MAC'
        wsdia['C1'] = 'CMTS'
        wsdia['D1'] = 'INTERFACE'
        wsdia['E1'] = 'NODE'
        wsdia['F1'] = 'IP CABLE'
        #wsdia['G1'] = 'IP CPE'
        #wsdia['H1'] = 'IP GW'
        wsdia['G1'] = 'TX'
        wsdia['H1'] = 'RX'
        wsdia['I1'] = 'SNR'
        wsdia['J1'] = 'ERRO LOG'


        editIntervalo = wsdia['A1']
        editIntervalo.font = Font(bold=True)
        editIntervalo.fill = PatternFill(fill_type='solid', start_color='FFC0C0C0', end_color='FFC0C0C0')
        wsdia['B1'].font = Font(bold=True)
        wsdia['B1'].fill = PatternFill(fill_type='solid', start_color='FFC0C0C0', end_color='FFC0C0C0')
        wsdia['C1'].font = Font(bold=True)
        wsdia['C1'].fill = PatternFill(fill_type='solid', start_color='FFC0C0C0', end_color='FFC0C0C0')
        wsdia['D1'].font = Font(bold=True)
        wsdia['D1'].fill = PatternFill(fill_type='solid', start_color='FFC0C0C0', end_color='FFC0C0C0')
        wsdia['E1'].font = Font(bold=True)
        wsdia['E1'].fill = PatternFill(fill_type='solid', start_color='FFC0C0C0', end_color='FFC0C0C0')
        wsdia['F1'].font = Font(bold=True)
        wsdia['F1'].fill = PatternFill(fill_type='solid', start_color='FFC0C0C0', end_color='FFC0C0C0')
        wsdia['G1'].font = Font(bold=True)
        wsdia['G1'].fill = PatternFill(fill_type='solid', start_color='FFC0C0C0', end_color='FFC0C0C0')
        wsdia['H1'].font = Font(bold=True)
        wsdia['H1'].fill = PatternFill(fill_type='solid', start_color='FFC0C0C0', end_color='FFC0C0C0')
        wsdia['I1'].font = Font(bold=True)
        wsdia['I1'].fill = PatternFill(fill_type='solid', start_color='FFC0C0C0', end_color='FFC0C0C0')
        wsdia['J1'].font = Font(bold=True)
        wsdia['J1'].fill = PatternFill(fill_type='solid', start_color='FFC0C0C0', end_color='FFC0C0C0')
        #wsdia['K1'].font = Font(bold=True)
        #wsdia['K1'].fill = PatternFill(fill_type='solid', start_color='FFC0C0C0', end_color='FFC0C0C0')
        #wsdia['L1'].font = Font(bold=True)
        #wsdia['L1'].fill = PatternFill(fill_type='solid', start_color='FFC0C0C0', end_color='FFC0C0C0')


        linha_iterada = 2
        coluna_iterada = 1
        ip_para_Logs = []

        texto = kwargs['listaMontada']

        with click.progressbar(texto, label='Consultando probes...:') as barra:

            for i in barra:

                listasPEV = i

                paginaPEV = 'http://pagina-de-consulta.com.br/cable/cable_docsis.php?strHost=S&mac5=' + listasPEV[0:2] + '&mac4=' + listasPEV[2:4] + '&mac3=' + listasPEV[4:6] + '&mac2=' + listasPEV[6:8] + '&mac1=' + listasPEV[8:10] + '&mac=' + listasPEV[10:12]

                requestPEV = requests.get(paginaPEV, auth=HTTPBasicAuth('login', 'senha'))

                requestPEV.raise_for_status()

                transformaSoup = bs4.BeautifulSoup(requestPEV.text, 'html.parser')

                elementoPage = transformaSoup.select('td')

                try:

                    vendorCMTS = elementoPage[9].getText().split(' ')
                    vendorCMTS = vendorCMTS[2][0:5]


                    if vendorCMTS == 'Cisco':

                        contrato = elementoPage[18].getText().split(' ')
                        contrato = contrato[2].strip()
                        cmts = elementoPage[7].getText().split(' ')
                        cmts = cmts[1].strip()
                        interface = elementoPage[8].getText().split(' ')
                        interface = interface[1].strip()
                        node = elementoPage[10].getText().split(' ')
                        node = node[1].strip()
                        ip = elementoPage[20].getText().split(' ')
                        ip = ip[6].strip()
                        #ipMib = subprocess.getoutput('snmpgetnext -v2c -c public '+ ip +' .1.3.6.1.2.1.4.20.1.1.127.0.0.1')
                        #ipMib = ipMib.split(' ')
                        #ipCpe = ipMib[3].strip()
                        #ipGw = subprocess.getoutput('snmpgetnext -v2c -On -c public '+ ip +' .1.3.6.1.2.1.4.24.7.1.10.1.4')
                        #ipGw = ipGw.split(' ')
                        #ipGw = ipGw[0].strip()
                        #ipGw = ipGw.split('.')
                        #ipTuple = (str(ipGw[24].strip()), str(ipGw[25].strip()), str(ipGw[26].strip()), str('0'))
                        #ipGw = '.'.join(ipTuple)
                        tx = elementoPage[27].getText().split(' ')
                        tx = tx[0].strip()
                        rx = elementoPage[29].getText().split(' ')
                        rx = rx[0].strip()
                        snr = elementoPage[31].getText().split(' ')
                        snr = snr[0].strip()
                        ip_para_Logs.append(ip)


                    if vendorCMTS == 'Arris':

                        contrato = elementoPage[16].getText().split(' ')
                        contrato = contrato[2].strip()
                        cmts = elementoPage[7].getText().split(' ')
                        cmts = cmts[1].strip()
                        interface = elementoPage[8].getText().split(' ')
                        interface = interface[3].strip()
                        node = elementoPage[10].getText().split(' ')
                        node = node[1].strip()
                        ip = elementoPage[18].getText().split(' ')
                        ip = ip[6].strip()
                        #ipMib = subprocess.getoutput('snmpgetnext -v2c -c public '+ ip +' .1.3.6.1.2.1.4.20.1.1.127.0.0.1')
                        #ipMib = ipMib.split(' ')
                        #ipCpe = ipMib[3].strip()
                        #ipGw = subprocess.getoutput('snmpgetnext -v2c -On -c public '+ ip +' .1.3.6.1.2.1.4.24.7.1.10.1.4')
                        #ipGw = ipGw.split(' ')
                        #ipGw = ipGw[0].strip()
                        #ipGw = ipGw.split('.')
                        #ipTuple = (str(ipGw[24].strip()), str(ipGw[25].strip()), str(ipGw[26].strip()), str('0'))
                        #ipGw = '.'.join(ipTuple)
                        tx = elementoPage[25].getText().split(' ')
                        tx = tx[0].strip()
                        rx = elementoPage[27].getText().split(' ')
                        rx = rx[0].strip()
                        snr = elementoPage[29].getText().split(' ')
                        snr = snr[0].strip()
                        ip_para_Logs.append(ip)


                    if vendorCMTS == ' ' or len(elementoPage) < 50:

                        contrato = 'OFFLINE'

                        cmts = ' '

                        interface = ' '

                        node = 'OFFLINE'

                        ip = ' '

                        ipCpe = ' '

                        ipGw = ' '

                        tx = '0'

                        rx = '0'

                        snr = '0'

                    if ip != ' ':

                        paginaLog = 'http://pagina-de-consulta.com.br/cable/log.php?ip_cable=' + str(ip)
                        requestLog = requests.get(paginaLog, auth=HTTPBasicAuth('login', 'senha'))
                        requestLog.raise_for_status()
                        logSoup = bs4.BeautifulSoup(requestLog.text, 'html.parser')

                        pegaRegex = re.compile(r'\s\w+\stime-out')

                        wsdia.cell(row=linha_iterada, column=coluna_iterada, value=contrato)
                        coluna_iterada += 1
                        wsdia.cell(row=linha_iterada, column=coluna_iterada, value=str(i))
                        coluna_iterada += 1
                        wsdia.cell(row=linha_iterada, column=coluna_iterada, value=cmts)
                        coluna_iterada += 1
                        wsdia.cell(row=linha_iterada, column=coluna_iterada, value=interface)
                        coluna_iterada += 1
                        wsdia.cell(row=linha_iterada, column=coluna_iterada, value=node)
                        coluna_iterada += 1
                        wsdia.cell(row=linha_iterada, column=coluna_iterada, value=ip)
                        coluna_iterada += 1
                        #wsdia.cell(row=linha_iterada, column=coluna_iterada, value=ipCpe)
                        #coluna_iterada += 1
                        #wsdia.cell(row=linha_iterada, column=coluna_iterada, value=ipGw)
                        #coluna_iterada += 1

                        wsdia.cell(row=linha_iterada, column=coluna_iterada, value=tx)

                        if float(tx) > 51:
                            wsdia.font = Font(color='FF1919', bold=True)
                        if float(tx) < 40:
                            wsdia.font = Font(color='FFD500', bold=True)

                        coluna_iterada += 1
                        wsdia.cell(row=linha_iterada, column=coluna_iterada, value=rx)

                        if float(rx) > 12:
                            wsdia.font = Font(color='FF1919', bold=True)
                        if float(rx) < -12:
                            wsdia.font = Font(color='FF1919', bold=True)

                        coluna_iterada += 1
                        wsdia.cell(row=linha_iterada, column=coluna_iterada, value=snr)

                        if float(snr) < 37:
                            wsdia.font = Font(color='FF1919', bold=True)

                        coluna_iterada += 1
                        wsdia.cell(row=linha_iterada, column=coluna_iterada, value=str(pegaRegex.findall(logSoup.text)))
                        coluna_iterada = 1


                        # ip_para_Logs.append(ip)

                        # print('MAC: ' + str(kwargs['listaMontada'][i]))
                        # print('CMTS: ' + elementoPage[7].getText())
                        # print('INTERFACE: ' + elementoPage[8].getText())
                        # print('Node: ' + elementoPage[10].getText())
                        # print('Contrato: ' + elementoPage[18].getText())
                        # print('IP: ' + elementoPage[20].getText())
                        # print('TX: ' + elementoPage[27].getText())
                        # print('RX: ' + elementoPage[29].getText())
                        # print('SNR: ' + elementoPage[31].getText())

                        # quantidades de td's com modem on: 61, off a pouco tempo: 21, off a muito tempo: 5. Descobri com a linha abaixo
                        # print('MAC: ' + str(kwargs['listaMontada'][i]) + '. Quantidade de tds: ' + str(len(elementoPage)))
                        # print(paginaPEV)

                        # print('MAC: ' + str(kwargs['listaMontada'][i]))
                        # print('OFFLINE')

                    linha_iterada += 1
                    click.clear()

                except:
                    pass #Se achar o erro, faz nada só avanca. to sem paciencia aff
                    #print("Oops! Erro " + sys.exc_info()[0] + "ocorreu.")
                    #print('')

        wb.save(caminho)
        print('Concluído!')
        print('')
        print('=' * 80)

        escolhaReset = input('Agora que terminamos, gostaria de resetar os logs dos equipamentos consultados? [S/N]')
        escolhaReset = escolhaReset.upper()

        if escolhaReset == 'N':
            print('OK, até mais!!!')
            exit()

        if escolhaReset == 'S':

            print('Vamos lá! isso pode demorar vários minutos...')

            for i in range(len(ip_para_Logs)):
                print('resetando IP..: ' + str(ip_para_Logs[i]))
                paginaResetLog = 'http://pagina-de-consulta.com.br/cable/ccm/logclear.php?IP=' + str(ip_para_Logs[i])
                requestResetLog = requests.get(paginaResetLog, auth=HTTPBasicAuth('login', 'senha'))
                requestResetLog.raise_for_status()
                if requestResetLog.status_code == 200:
                    print('Sucesso!')
                else:
                    print('Erro ao resetar ip: ' + str(ip_para_Logs[i]) + '. Cód. do erro: ' + str(
                        requestResetLog.status_code))
            exit()

        if escolhaReset != 'N' or escolhaReset != 'S':
            print(
                'Não entendi qual a opção escolhida, então estarei encerrando por aqui de qualquer maneira, até mais!!!')




    else:
        print(
            'Não foi possível achar o arquivo \"consultados.xlsx\". Verifique se o arquivo existe ou se está reservado para outro usuário')
