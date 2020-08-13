#!/home/user/.virtualenvs/k36/bin/python
# -*- coding: utf-8 -*-
# Author: Valdemir Bezerra

import os, sys
import subprocess
import openpyxl
from openpyxl.styles import PatternFill
import re
from datetime import date
import time
import shelve

from email.mime.multipart import MIMEMultipart
from email.mime.multipart import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import smtplib

print('''

  _______   _____     ____    _    _    _____   _    _   _____    _    _   _______ 
 |__   __| |  __ \   / __ \  | |  | |  / ____| | |  | | |  __ \  | |  | | |__   __|
    | |    | |__) | | |  | | | |  | | | |  __  | |__| | | |__) | | |  | |    | |   
    | |    |  _  /  | |  | | | |  | | | | |_ | |  __  | |  ___/  | |  | |    | |   
    | |    | | \ \  | |__| | | |__| | | |__| | | |  | | | |      | |__| |    | |   
    |_|    |_|  \_\  \____/   \____/   \_____| |_|  |_| |_|       \____/     |_|   
                                                                                   
                                                                                   
                        Powered on PYTHON!!!
                        Developer: VALDEMIR

''')

if not os.geteuid() == 0:
    mensagemErro = "\nErro:\n Voce deve estar logado como root no prompt de comando para usar essa aplicaçao.\nUsuario ativo: " + \
                   os.environ['USER']
    sys.exit(mensagemErro)

print("Acesso root concedido! Dando inicio a coleta de troughput...")


def salvaExcel(**kwargs):
    hoje = date.today()

    wb = openpyxl.load_workbook('/home/user/Documentos/SCRIPTS/troughput/equissel/salvatroughput.xlsx',
                                data_only=True)
    wb.sheetnames
    sheet = wb[kwargs['roteador']]

    ultimaColunaUsada = sheet.max_column
    ultimaLinhaUsada = sheet.max_row + 1

    novaColuna = ultimaColunaUsada + 1

    sheet.cell(row=1, column=novaColuna).value = "data-in-" + str(hoje.day) + "/" + str(hoje.month) + "/" + str(
        hoje.year)[2:4]
    sheet.cell(row=1, column=novaColuna + 1).value = "data-out-" + str(hoje.day) + "/" + str(
        hoje.month) + "/" + str(hoje.year)[2:4]

    linhaInicialIn = 2
    linhaInicialOut = 2
    dividendo = 0

    scoreLinhas = kwargs['escore'] * 2

    for i in kwargs['listaT']:
        if dividendo % 2 == 0:
            sheet.cell(row=linhaInicialIn, column=novaColuna).value = i
            linhaInicialIn += 1
        else:
            sheet.cell(row=linhaInicialOut, column=novaColuna + 1).value = i
            linhaInicialOut += 1

        dividendo += 1

    wb.save('/home/user/Documentos/SCRIPTS/troughput/equissel/salvatroughput.xlsx')
    print(' ')
    print('Arquivo excel atualizado com sucesso!')
    print(' ')


def calculaMediaExcel():
    wb = openpyxl.load_workbook('/home/user/Documentos/SCRIPTS/troughput/equissel/salvatroughput.xlsx')
    wb.sheetnames

    roteadores = os.listdir(os.path.join('/home/user/Documentos/SCRIPTS/troughput/rt/'))

    for roteador in roteadores:

        #sheet = wb.get_sheet_by_name(roteador)
        sheet = wb[roteador]
        hoje = date.today()

        ultimaColunaUsada = sheet.max_column
        novaColuna = ultimaColunaUsada + 1
        ultimaLinhaUsada = sheet.max_row

        sheet.cell(row=1, column=novaColuna).fill = PatternFill(fill_type='solid', start_color='EE1111',
                                                                end_color='EE1111')

        sheet.cell(row=1, column=novaColuna + 1).fill = PatternFill(fill_type='solid', start_color='EE1111',
                                                                    end_color='EE1111')

        sheet.cell(row=1, column=novaColuna).value = "MEDIA-in-" + str(hoje.day) + "/" + str(hoje.month) + "/" + str(
            hoje.year)[2:4]
        sheet.cell(row=1, column=novaColuna + 1).value = "MEDIA-out-" + str(hoje.day) + "/" + str(
            hoje.month) + "/" + str(hoje.year)[2:4]

        arquivoSRC = shelve.open(os.path.join('/home/user/Documentos/SCRIPTS/troughput/scr/', 'score'))
        score = arquivoSRC['scr']

        scoreLinhas = score * 2

        arquivoSRC.close()

        for linhaSelecionada in range(2, ultimaLinhaUsada + 1):

            coordenadasLinhasIn = " "
            coordenadasLinhasOut = " "
            dividendo = 0

            for linhaMedia in range(scoreLinhas, 0, -1):

                if dividendo % 2 == 0:
                    coordenadasLinhasIn += sheet.cell(row=linhaSelecionada,
                                                      column=novaColuna - linhaMedia).coordinate + "+"

                else:
                    coordenadasLinhasOut += sheet.cell(row=linhaSelecionada,
                                                       column=novaColuna - linhaMedia).coordinate + "+"

                dividendo += 1

            sheet.cell(row=linhaSelecionada, column=novaColuna).value = '=((' + coordenadasLinhasIn[: -1] + ')/' + str(
                score) + ')/1000000000'

            sheet.cell(row=linhaSelecionada, column=novaColuna + 1).value = '=((' + coordenadasLinhasOut[
                                                                                    : -1] + ')/' + str(
                score) + ')/1000000000'

        wb.save('/home/user/Documentos/SCRIPTS/troughput/equissel/salvatroughput.xlsx')


def salvaEscore(somaEscore):
    if os.path.exists(os.path.join('/home/user/Documentos/SCRIPTS/troughput/scr/', 'score')) == True:

        if somaEscore == 0:

            arquivoSRC = shelve.open(os.path.join('/home/user/Documentos/SCRIPTS/troughput/scr/', 'score'))

            valor = arquivoSRC['scr']

            soma = int(valor) + 1

            arquivoSRC['scr'] = soma

            arquivoSRC.close()

        elif somaEscore == 1:
            arquivoSRC = shelve.open(os.path.join('/home/user/Documentos/SCRIPTS/troughput/scr/', 'score'))
            scr = 0
            arquivoSRC['scr'] = scr
            arquivoSRC.close()

    else:

        arquivoSRC = shelve.open(os.path.join('/home/user/Documentos/SCRIPTS/troughput/scr/', 'score'))
        scr = 0
        arquivoSRC['scr'] = scr
        arquivoSRC.close()


def exec_rotina(*args):
    hoje = date.today()

    if hoje.weekday() == 4:
        salvaEscore(0)

        arquivoSRC = shelve.open(os.path.join('/home/user/Documentos/SCRIPTS/troughput/scr/', 'score'))
        valor = arquivoSRC['scr']
        arquivoSRC.close()

        for roteador in args:
            processo = subprocess.run(os.path.join('/home/user/Documentos/SCRIPTS/troughput/rt/', roteador),
                                      shell=True)

            pegaregex = re.compile(r'(\d+)\sbits/sec')

            arquivoSalvo = roteador
            retornoRoteador = arquivoSalvo.split(".")

            conteudoArquivo = open(
                os.path.join('/home/user/Documentos/SCRIPTS/troughput/tmp/', retornoRoteador[0]))
            stringArquivo = conteudoArquivo.read()

            listaTroughput = pegaregex.findall(stringArquivo)

            salvaExcel(roteador=arquivoSalvo, escore=valor, listaT=listaTroughput)

        print("Hoje e sexta-feira, dia de maldade e de calcular o troughput semanal!")
        print("Calculando...")
        calculaMediaExcel()

        salvaEscore(1)

        print('Estamos preparando tudo para enviar o resultado semanal, por favor aguarde...')

        time.sleep(6)

        mandaEmail()

    else:

        salvaEscore(0)

        for roteador in args:
            processo = subprocess.run(os.path.join('/home/user/Documentos/SCRIPTS/troughput/rt/', roteador),
                                      shell=True)

            pegaregex = re.compile(r'(\d+)\sbits/sec')

            arquivoSalvo = roteador
            retornoRoteador = arquivoSalvo.split(".")

            conteudoArquivo = open(
                os.path.join('/home/user/Documentos/SCRIPTS/troughput/tmp/', retornoRoteador[0]))
            stringArquivo = conteudoArquivo.read()

            listaTroughput = pegaregex.findall(stringArquivo)

            salvaExcel(roteador=arquivoSalvo, escore=0, listaT=listaTroughput)


def mandaEmail():
    print(' ')

    hoje = date.today()
    msg = MIMEMultipart()

    msg['Subject'] = "Check Troughput - " + str(hoje.day) + "/" + str(hoje.month) + "/" + str(hoje.year)
    msg.attach(MIMEText('Prezados, segue check de troughput'))

    part = MIMEBase('application', 'octect-stream')
    part.set_payload(
        open(os.path.join('/home/user/Documentos/SCRIPTS/troughput/equissel/', 'salvatroughput.xlsx'),
             'rb').read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment; filename="salvatroughput.xlsx"')

    msg.attach(part)

    smtpObj = smtplib.SMTP('smtp.gmail.com', 587)

    print('Aguardando resposta do server...')

    print('')

    print(smtpObj.ehlo())  # resposta do server

    print('')

    print('Aguardando status tls do server...')

    print('')

    print(smtpObj.starttls())  # status do TLS

    print('')

    print('Enviando credenciais de acesso...')

    print(smtpObj.login('email-origem@gmail.com', 'senhaemailenvioautomatico'))  # login

    print('')

    print('OK! Mandando email...')

    smtpObj.sendmail('email-origem@gmail.com', ['email-destino@gmail.com.br', 'email-destino2@gmail.com.br'],
                     msg.as_string())

    print('Fechando conexao...')

    print(smtpObj.quit())

    print('')

    print('Email enviado!')


print('Listando roteadores para consulta...')

roteadores = os.listdir(os.path.join('/home/user/Documentos/SCRIPTS/troughput/rt/'))

exec_rotina(*roteadores)

for i in roteadores:

    arquivoSalvo = i
    retornoRoteador = arquivoSalvo.split(".")

    os.remove(os.path.join('/home/user/Documentos/SCRIPTS/troughput/tmp/', retornoRoteador[0]))

print("")
print("Coleta de troughput realizada com sucesso!")
